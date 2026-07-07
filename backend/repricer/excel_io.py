"""Чтение файла проценки и запись результата.

Стратегия: входной файл открывается дважды —
- data_only=True  → кэшированные значения формул (уценка в J посчитана VLOOKUP'ом);
- data_only=False → рабочая книга для записи (сохраняет форматирование и прочие листы).

В выходном файле формулы колонки J (VLOOKUP на внешнюю книгу) замораживаются
в значения: внешняя ссылка при открытии копии всё равно битая и дала бы #REF!.
Колонка L пишется формулой =K{row}-F{row}. Всё остальное — без изменений.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Union

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from repricer.core import PriceResult, Rules, Status, compute_price

REVIEW_SHEET = "На разбор"
STATUS_HEADER = "Статус"

# Обязательные колонки (резолв только по заголовку; для склада есть фолбэк ниже)
REQUIRED_COLUMNS = {
    "артикул": "article",
    "цена": "cost",
    "min цена": "min_price",
    "склад/цех": "warehouse",
}
# Опциональные: есть в 13-колоночном формате, могут отсутствовать в 8-колоночном.
# Уценка из файла имеет приоритет над любым внешним справочником: если колонка
# присутствует — берём её значения, внешние источники не опрашиваются.
OPTIONAL_COLUMNS = {
    "#": "num",
    "бренд": "brand",
    "номенклатура": "name",
    "количество": "qty",
    "цена согл с уценкой": "markdown",
    "уценка согл.": "markdown",   # вариант заголовка из файла от 06.07 с уценкой
    "поставщик": "supplier",
    "новая цена": "new_price",
    "дельта ц.нов-себес": "delta",
    # «Нов цена-4%» из того же файла — собственный расчёт оператора, НЕ маппим:
    # колонку не трогаем, наши результаты дописываются отдельными колонками
}
# Ключи, которые всегда присутствуют в values строки (отсутствующая колонка → None)
ROW_VALUE_KEYS = (
    "num", "article", "brand", "name", "qty", "cost",
    "min_price", "supplier", "markdown", "warehouse",
)
HEADER_SEARCH_ROWS = 5
CONTENT_SAMPLE_ROWS = 50


def _norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower()) if value is not None else ""


@dataclass
class RowOutcome:
    row: int                 # номер строки листа Excel
    result: PriceResult
    values: dict             # исходные значения строки (для листа «На разбор»)


@dataclass
class RepriceReport:
    sheet: str
    total: int = 0
    by_status: Dict[str, int] = field(default_factory=dict)
    review_rows: List[RowOutcome] = field(default_factory=list)
    all_rows: List[RowOutcome] = field(default_factory=list)

    @property
    def review_count(self) -> int:
        return len(self.review_rows)


def find_header_row(ws: Worksheet) -> Optional[int]:
    """Строка заголовка: в первых 5 строках есть ячейки «Артикул» и «Цена».

    Позиции колонок не хардкодим — только содержимое заголовков.
    """
    for row in range(1, min(HEADER_SEARCH_ROWS, ws.max_row) + 1):
        headers = {_norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if "артикул" in headers and "цена" in headers:
            return row
    return None


def _column_signature(ws: Worksheet, header_row: int, col: int) -> Optional[str]:
    """Тип содержимого беззаголовочной колонки: 'text' / 'number' / None (пусто/смешано)."""
    texts = numbers = total = 0
    for row in range(header_row + 1, min(header_row + CONTENT_SAMPLE_ROWS, ws.max_row) + 1):
        value = ws.cell(row, col).value
        if value is None or value == "":
            continue
        total += 1
        if isinstance(value, str):
            texts += 1
        elif isinstance(value, (int, float)):
            numbers += 1
    if total == 0:
        return None
    if texts / total >= 0.9:
        return "text"
    if numbers / total >= 0.9:
        return "number"
    return None


def map_columns(ws: Worksheet, header_row: int) -> Dict[str, int]:
    """Заголовок → индекс колонки. Бросает ValueError, если обязательной колонки нет.

    Фолбэк для «Склад/Цех»: в 8-колоночном формате у колонки склада нет заголовка.
    Если заголовок не найден, ищем среди беззаголовочных колонок единственную
    текстовую — это резолв по содержимому, а не по позиции. Неоднозначность — ошибка.
    Аналогично «#»: единственная беззаголовочная числовая колонка.
    """
    found: Dict[str, int] = {}
    unheaded: List[int] = []
    for col in range(1, ws.max_column + 1):
        header = _norm(ws.cell(header_row, col).value)
        if not header:
            unheaded.append(col)
            continue
        if header in REQUIRED_COLUMNS and REQUIRED_COLUMNS[header] not in found:
            found[REQUIRED_COLUMNS[header]] = col
        if header in OPTIONAL_COLUMNS and OPTIONAL_COLUMNS[header] not in found:
            found[OPTIONAL_COLUMNS[header]] = col

    if "warehouse" not in found:
        text_cols = [c for c in unheaded if _column_signature(ws, header_row, c) == "text"]
        if len(text_cols) == 1:
            found["warehouse"] = text_cols[0]
        elif len(text_cols) > 1:
            letters = ", ".join(get_column_letter(c) for c in text_cols)
            raise ValueError(
                f"Не удалось определить колонку «Склад/Цех»: заголовка нет, "
                f"а текстовых колонок без заголовка несколько ({letters})"
            )

    if "num" not in found:
        num_cols = [c for c in unheaded if _column_signature(ws, header_row, c) == "number"]
        if len(num_cols) == 1:
            found["num"] = num_cols[0]

    missing = [h for h, key in REQUIRED_COLUMNS.items() if key not in found]
    if missing:
        raise ValueError(f"Во входном файле не найдены колонки: {', '.join(missing)}")
    return found


def _pick_sheet(wb) -> str:
    for name in wb.sheetnames:
        if find_header_row(wb[name]) is not None:
            return name
    raise ValueError(
        "Не найден лист проценки: ни на одном листе нет строки заголовка с '#' и 'Артикул'"
    )


def reprice_file(
    input_path: Union[str, Path],
    output_path: Union[str, Path],
    rules: Rules,
) -> RepriceReport:
    input_path, output_path = Path(input_path), Path(output_path)

    wb_values = openpyxl.load_workbook(input_path, data_only=True)
    wb_out = openpyxl.load_workbook(input_path, data_only=False)

    sheet_name = _pick_sheet(wb_values)
    ws_vals = wb_values[sheet_name]
    ws_out = wb_out[sheet_name]

    header_row = find_header_row(ws_vals)
    cols = map_columns(ws_vals, header_row)
    report = RepriceReport(sheet=sheet_name)

    c = cols  # короткий алиас

    # Выходные колонки: если их нет во входном формате — дописываем в конец
    next_free = ws_vals.max_column + 1
    for key, header in (("new_price", "Новая цена"), ("delta", "дельта ц.нов-себес")):
        if key not in c:
            c[key] = next_free
            ws_out.cell(header_row, c[key], header).font = Font(bold=True)
            next_free += 1
    status_col = next_free
    ws_out.cell(header_row, status_col, STATUS_HEADER).font = Font(bold=True)

    k_letter = get_column_letter(c["new_price"])
    f_letter = get_column_letter(c["cost"])

    for row in range(header_row + 1, ws_vals.max_row + 1):
        vals = {key: ws_vals.cell(row, col).value for key, col in cols.items() if key in ROW_VALUE_KEYS}
        for key in ROW_VALUE_KEYS:
            vals.setdefault(key, None)
        # пропускаем пустые строки (ни артикула, ни цен)
        if all(vals[k] in (None, "") for k in ("article", "cost", "min_price")):
            continue

        report.total += 1
        result = compute_price(vals["cost"], vals["min_price"], vals["markdown"], rules)
        report.by_status[result.status] = report.by_status.get(result.status, 0) + 1

        if result.new_price is not None:
            price = result.new_price
            # целочисленный шаг → пишем int, чтобы в Excel не было "1388.0"
            ws_out.cell(row, c["new_price"], int(price) if price == price.to_integral_value() else float(price))
        # цена None (НЕТ MIN ЦЕНЫ / ОШИБКА ДАННЫХ) — «Новую цену» не трогаем

        ws_out.cell(row, c["delta"], f"={k_letter}{row}-{f_letter}{row}")
        ws_out.cell(row, status_col, result.status)

        # Замораживаем формулу уценки (VLOOKUP на внешний файл) в значение
        if "markdown" in c:
            md_cell = ws_out.cell(row, c["markdown"])
            if isinstance(md_cell.value, str) and md_cell.value.startswith("="):
                cached = vals["markdown"]
                md_cell.value = None if isinstance(cached, str) else cached

        item = RowOutcome(row=row, result=result, values=vals)
        report.all_rows.append(item)
        if result.for_review:
            report.review_rows.append(item)

    _write_review_sheet(wb_out, report)
    wb_out.save(output_path)
    return report


REVIEW_HEADERS = [
    "Строка файла", "#", "Артикул", "Бренд", "Номенклатура", "Количество",
    "Себестоимость", "Min Цена", "Цена согл с уценкой", "Новая цена", "Статус", "Склад/Цех",
]


def _write_review_sheet(wb, report: RepriceReport) -> None:
    if REVIEW_SHEET in wb.sheetnames:
        del wb[REVIEW_SHEET]
    ws = wb.create_sheet(REVIEW_SHEET)
    for col, header in enumerate(REVIEW_HEADERS, start=1):
        ws.cell(1, col, header).font = Font(bold=True)

    for i, item in enumerate(report.review_rows, start=2):
        v, r = item.values, item.result
        markdown = None if isinstance(v["markdown"], str) else v["markdown"]
        new_price = r.new_price
        if isinstance(new_price, Decimal):
            new_price = int(new_price) if new_price == new_price.to_integral_value() else float(new_price)
        row_values = [
            item.row, v["num"], v["article"], v["brand"], v["name"], v["qty"],
            v["cost"], v["min_price"], markdown, new_price, r.status, v["warehouse"],
        ]
        for col, value in enumerate(row_values, start=1):
            ws.cell(i, col, value)

    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["L"].width = 45
