from pathlib import Path

import openpyxl
import pytest

from repricer.core import Rules
from repricer.excel_io import REVIEW_SHEET, find_header_row, map_columns, reprice_file

HEADERS = ["#", "Артикул", "Бренд", "Номенклатура", "Количество", "Цена", "Min Цена",
           None, "Поставщик", "Цена согл с  уценкой", "Новая цена", "дельта ц.нов-себес", "Склад/Цех"]


def make_input(path: Path, rows):
    """Собрать входной файл в формате реальной проценки: пустая 1-я строка, заголовок во 2-й."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Проценка"
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(2, col, header)
    for i, row in enumerate(rows, start=3):
        for col, value in enumerate(row, start=1):
            ws.cell(i, col, value)
    wb.save(path)
    return path


ROWS = [
    # #, Артикул, Бренд, Номенклатура, Кол, Цена(себест), Min, H, Поставщик, Уценка, Новая, дельта, Склад
    [1, "9046709050", "TOYOTA", "Клипса", 10, 500, 1000, None, "site.ru", None, 111, "=K3-F3", "Склад А"],
    [2, "0012 33", "VAG", "Насадка", 2, 2000, 1000, None, "site.ru", 500, 111, "=K4-F4", "Склад Б"],
    [3, "ABC", "FORD", "Зеркало", 1, 5000, 1000, None, "site.ru", None, 111, "=K5-F5", "Склад А"],
    [4, "DEF", "KIA", "Фара", 1, 100, None, None, None, None, 111, "=K6-F6", "Склад Б"],
]


@pytest.fixture
def result(tmp_path):
    inp = make_input(tmp_path / "in.xlsx", ROWS)
    out = tmp_path / "out.xlsx"
    report = reprice_file(inp, out, Rules())
    return report, openpyxl.load_workbook(out)


def test_summary(result):
    report, _ = result
    assert report.total == 4
    assert report.by_status["OK"] == 1
    assert report.by_status["УЦЕНКА+5%"] == 1
    assert report.by_status["КОНФЛИКТ"] == 1
    assert report.by_status["НЕТ MIN ЦЕНЫ"] == 1
    assert report.review_count == 2


def test_prices_and_statuses(result):
    _, wb = result
    ws = wb["Проценка"]
    assert ws.cell(3, 11).value == 960          # OK: 1000*0.96
    assert ws.cell(4, 11).value == 525          # УЦЕНКА: 500*1.05
    assert ws.cell(5, 11).value == 5200         # КОНФЛИКТ: 5000*1.04 > Min
    assert ws.cell(6, 11).value == 111          # НЕТ MIN — цена не изменена
    assert [ws.cell(r, 14).value for r in range(3, 7)] == ["OK", "УЦЕНКА+5%", "КОНФЛИКТ", "НЕТ MIN ЦЕНЫ"]
    assert ws.cell(2, 14).value == "Статус"


def test_delta_formula(result):
    _, wb = result
    ws = wb["Проценка"]
    for r in range(3, 7):
        assert ws.cell(r, 12).value == f"=K{r}-F{r}"


def test_untouched_columns(result):
    _, wb = result
    ws = wb["Проценка"]
    for i, row in enumerate(ROWS, start=3):
        assert ws.cell(i, 2).value == row[1]    # артикул как строка, с пробелами
        assert ws.cell(i, 13).value == row[12]  # Склад/Цех 1:1
        assert ws.cell(i, 5).value == row[4]    # количество


def test_review_sheet(result):
    _, wb = result
    ws = wb[REVIEW_SHEET]
    assert ws.cell(1, 1).value == "Строка файла"
    got = {(ws.cell(r, 1).value, ws.cell(r, 11).value) for r in (2, 3)}
    assert got == {(5, "КОНФЛИКТ"), (6, "НЕТ MIN ЦЕНЫ")}
    # артикул и склад продублированы для удобства разбора
    assert ws.cell(3, 3).value == "DEF"
    assert ws.cell(3, 12).value == "Склад Б"


def test_markdown_formula_frozen(tmp_path):
    rows = [[1, "X1", "VAG", "Деталь", 1, 500, 520, None, "site.ru",
             "=VLOOKUP(B3,[1]Лист2!$A:$J,10,0)", 111, "=K3-F3", "Склад А"]]
    inp = make_input(tmp_path / "in.xlsx", rows)
    out = tmp_path / "out.xlsx"
    # data_only=True вернёт None для незакешированной формулы → уценки нет → СЕБЕСТ
    report = reprice_file(inp, out, Rules())
    ws = openpyxl.load_workbook(out)["Проценка"]
    assert ws.cell(3, 10).value is None  # формула с внешней ссылкой заморожена
    assert report.by_status == {"СЕБЕСТ+4%": 1}


def test_header_row_not_hardcoded(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(5, col, header)  # заголовок в 5-й строке
    ws.cell(6, 1, 1); ws.cell(6, 2, "ART"); ws.cell(6, 4, "Деталь")
    ws.cell(6, 6, 500); ws.cell(6, 7, 1000); ws.cell(6, 13, "Склад")
    inp = tmp_path / "in.xlsx"
    wb.save(inp)
    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())
    assert report.total == 1
    assert report.by_status == {"OK": 1}


def test_missing_column_raises(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(HEADERS)
    headers[6] = "Другое"  # убрали Min Цена
    for col, header in enumerate(headers, start=1):
        ws.cell(2, col, header)
    inp = tmp_path / "in.xlsx"
    wb.save(inp)
    with pytest.raises(ValueError, match="min цена"):
        reprice_file(inp, tmp_path / "out.xlsx", Rules())


def test_no_header_anywhere_raises(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.cell(1, 1, "просто данные")
    inp = tmp_path / "in.xlsx"
    wb.save(inp)
    with pytest.raises(ValueError, match="Не найден лист"):
        reprice_file(inp, tmp_path / "out.xlsx", Rules())


def test_empty_rows_skipped(tmp_path):
    rows = list(ROWS) + [[None] * 13]
    inp = make_input(tmp_path / "in.xlsx", rows)
    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())
    assert report.total == 4
