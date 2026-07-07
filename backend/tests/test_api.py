import base64
from io import BytesIO
from zipfile import ZipFile

import openpyxl
from fastapi.testclient import TestClient

from app.main import app
from tests.test_excel_io import make_input

client = TestClient(app)


def test_health():
    response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def _post_xlsx(path, data=None):
    with path.open("rb") as fh:
        return client.post(
            "/api/reprice",
            data=data or {},
            files={
                "file": (
                    path.name,
                    fh.read(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )


def test_reprice_happy_path(tmp_path):
    inp = make_input(
        tmp_path / "in.xlsx",
        [
            [1, "9046709050", "TOYOTA", "Клипса", 10, 500, 1000, None, "site.ru", None, 111, "=K3-F3", "Склад А"],
            [2, "ABC", "FORD", "Зеркало", 1, 5000, 1000, None, "shop.avtogut.ru", None, 111, "=K4-F4", "Склад Б"],
        ],
    )

    response = _post_xlsx(inp)

    assert response.status_code == 200
    data = response.json()
    assert data["summary"] == {
        "sheet": "Проценка",
        "total": 2,
        "by_status": {"OK": 1, "КОНФЛИКТ": 1},
        "review_count": 1,
        "warnings": [],
        "params": {
            "discount": 0.04,
            "markdown_markup": 0.05,
            "cost_markup": 0.04,
            "rounding": "1",
        },
    }
    assert data["output_filename"] == "in_repriced.xlsx"
    assert len(data["rows"]) == 2
    assert data["rows"][0] == {
        "row": 3,
        "num": 1,
        "article": "9046709050",
        "brand": "TOYOTA",
        "name": "Клипса",
        "qty": 10,
        "cost": 500,
        "min_price": 1000,
        "supplier": "site.ru",
        "markdown": None,
        "new_price": 960,
        "delta": 460,
        "status": "OK",
        "for_review": False,
        "warehouse": "Склад А",
    }
    assert data["rows"][1]["supplier"] == "shop.avtogut.ru"
    assert data["rows"][1]["status"] == "КОНФЛИКТ"
    decoded = base64.b64decode(data["output_xlsx_base64"])
    wb = openpyxl.load_workbook(BytesIO(decoded))
    assert "Проценка" in wb.sheetnames
    assert wb["Проценка"].cell(3, 11).value == 960


def test_reprice_custom_params_change_price(tmp_path):
    inp = make_input(
        tmp_path / "custom.xlsx",
        [[1, "A1", "TOYOTA", "Клипса", 1, 100, 1000, None, "site.ru", None, 111, "=K3-F3", "Склад А"]],
    )

    response = _post_xlsx(inp, data={"discount": "0.05"})

    assert response.status_code == 200
    data = response.json()
    assert data["summary"]["params"]["discount"] == 0.05
    assert data["rows"][0]["new_price"] == 950
    assert data["rows"][0]["delta"] == 850


def test_reprice_missing_required_column_returns_400(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Проценка"
    headers = [
        "#", "Артикул", "Бренд", "Номенклатура", "Количество", "Цена",
        "Другое", None, "Поставщик", "Цена согл с  уценкой",
        "Новая цена", "дельта ц.нов-себес", "Склад/Цех",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(2, col, header)
    inp = tmp_path / "bad.xlsx"
    wb.save(inp)

    response = _post_xlsx(inp)

    assert response.status_code == 400
    assert "min цена" in response.json()["detail"]


def test_reprice_invalid_rounding_returns_400(tmp_path):
    inp = make_input(
        tmp_path / "rounding.xlsx",
        [[1, "A1", "TOYOTA", "Клипса", 1, 100, 1000, None, None, None, 111, "=K3-F3", "Склад А"]],
    )

    response = _post_xlsx(inp, data={"rounding": "5"})

    assert response.status_code == 400
    assert "rounding" in response.json()["detail"]


def test_reprice_non_xlsx_returns_400(tmp_path):
    inp = tmp_path / "in.txt"
    inp.write_text("not an excel file", encoding="utf-8")

    response = _post_xlsx(inp)

    assert response.status_code == 400
    assert response.json()["detail"] == "Поддерживаются только .xlsx файлы"


def test_reprice_invalid_zip_xlsx_returns_400(tmp_path):
    inp = tmp_path / "not_excel.xlsx"
    with ZipFile(inp, "w") as zf:
        zf.writestr("data.txt", "not an excel workbook")

    response = _post_xlsx(inp)

    assert response.status_code == 400
