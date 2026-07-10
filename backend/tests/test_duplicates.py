"""Дубли артикула (разные партии): расчёт от минимальной Min цены группы."""
import openpyxl
import pytest

from repricer.core import Rules
from repricer.excel_io import reprice_file
from tests.test_excel_io import make_input


def rows_with(article_min_pairs):
    """Строки 13-колоночного формата: (артикул, себест, min) → полная строка."""
    out = []
    for i, (article, cost, min_price) in enumerate(article_min_pairs, start=1):
        out.append([i, article, "BRAND", f"Деталь {i}", 1, cost, min_price,
                    None, "site.ru", None, None, None, "Склад А"])
    return out


def test_duplicates_use_group_min(tmp_path):
    # две партии одного артикула: Min 1000 и 800 → обе считаются от 800
    inp = make_input(tmp_path / "in.xlsx", rows_with([
        ("DUP1", 100, 1000),
        ("DUP1", 200, 800),
        ("UNIQ", 100, 1000),
    ]))
    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())

    ws = openpyxl.load_workbook(tmp_path / "out.xlsx")["Проценка"]
    assert ws.cell(3, 11).value == 768   # 800×0.96, а не 960
    assert ws.cell(4, 11).value == 768
    assert ws.cell(5, 11).value == 960   # уникальный артикул — от собственной Min

    # подмена зафиксирована только у строки с чужой Min
    assert report.all_rows[0].effective_min == 800
    assert report.all_rows[1].effective_min is None
    assert report.all_rows[2].effective_min is None


def test_duplicates_warning(tmp_path):
    inp = make_input(tmp_path / "in.xlsx", rows_with([
        ("DUP1", 100, 1000),
        ("DUP1", 100, 800),
    ]))
    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())
    assert len(report.warnings) == 1
    assert "партий" in report.warnings[0]


def test_no_duplicates_no_warning(tmp_path):
    inp = make_input(tmp_path / "in.xlsx", rows_with([
        ("A1", 100, 1000),
        ("A2", 100, 800),
    ]))
    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())
    assert report.warnings == []


def test_duplicate_rescues_missing_min(tmp_path):
    # у одной партии Min пустая — наследует групповую, а не «НЕТ MIN ЦЕНЫ»
    inp = make_input(tmp_path / "in.xlsx", rows_with([
        ("DUP1", 100, None),
        ("DUP1", 100, 900),
    ]))
    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())
    assert report.by_status == {"OK": 2}
    assert report.all_rows[0].effective_min == 900


def test_single_row_missing_min_still_flagged(tmp_path):
    inp = make_input(tmp_path / "in.xlsx", rows_with([("SOLO", 100, None)]))
    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())
    assert report.by_status == {"НЕТ MIN ЦЕНЫ": 1}


def test_duplicates_same_min_no_substitution(tmp_path):
    inp = make_input(tmp_path / "in.xlsx", rows_with([
        ("DUP1", 100, 1000),
        ("DUP1", 200, 1000),
    ]))
    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())
    assert report.warnings == []
    assert all(item.effective_min is None for item in report.all_rows)


def test_int_and_str_article_same_key(tmp_path):
    # артикул 12345 (int) и «12345» (str с пробелом) — одна группа
    inp = make_input(tmp_path / "in.xlsx", rows_with([
        (12345, 100, 1000),
        ("12345 ", 100, 700),
    ]))
    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())
    ws = openpyxl.load_workbook(tmp_path / "out.xlsx")["Проценка"]
    assert ws.cell(3, 11).value == 672   # 700×0.96
    assert report.all_rows[0].effective_min == 700
