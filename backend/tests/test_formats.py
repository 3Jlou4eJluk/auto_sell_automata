"""Единый парсер двух форматов проценки: 13-колоночного и 8-колоночного."""
from pathlib import Path

import openpyxl
import pytest

from repricer.core import Rules
from repricer.excel_io import reprice_file

FIXTURES = Path(__file__).parent / "fixtures"

# 8-колоночный формат от 06.07: «№» (A) и «Склад/Цех» (H) БЕЗ заголовков,
# нет уценки, поставщика, «Новой цены» и «дельты»
HEADERS_8COL = [None, "Артикул", "   Номенклатура", "Бренд", "Цена", "Min Цена", "Количество", None]


def make_8col(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Проценка"
    for col, header in enumerate(HEADERS_8COL, start=1):
        ws.cell(2, col, header)
    for i, row in enumerate(rows, start=3):
        for col, value in enumerate(row, start=1):
            ws.cell(i, col, value)
    wb.save(path)
    return path


ROWS_8COL = [
    # №, Артикул, Номенклатура, Бренд, Цена(себест), Min, Кол-во, Склад(без заголовка)
    [1, "9046709050", "Клипса", "TOYOTA", 500, 1000, 10, "Склад Челны"],
    [2, "ABC", "Зеркало", "FORD", 5000, 1000, 1, "Склад Уфа"],
]


def test_8col_parses_and_appends_output_columns(tmp_path):
    inp = make_8col(tmp_path / "in.xlsx", ROWS_8COL)
    out = tmp_path / "out.xlsx"
    report = reprice_file(inp, out, Rules())

    assert report.total == 2
    assert report.by_status == {"OK": 1, "КОНФЛИКТ": 1}

    ws = openpyxl.load_workbook(out)["Проценка"]
    # выходные колонки дописаны в конец: I=Новая цена, J=дельта, K=Статус
    assert ws.cell(2, 9).value == "Новая цена"
    assert ws.cell(2, 10).value == "дельта ц.нов-себес"
    assert ws.cell(2, 11).value == "Статус"
    assert ws.cell(3, 9).value == 960          # 1000*0.96
    assert ws.cell(3, 10).value == "=I3-E3"    # дельта от колонки Цена (E)
    assert ws.cell(3, 11).value == "OK"
    # склад и артикул не тронуты
    assert ws.cell(3, 8).value == "Склад Челны"
    assert ws.cell(4, 8).value == "Склад Уфа"


def test_8col_warehouse_resolved_without_header(tmp_path):
    inp = make_8col(tmp_path / "in.xlsx", ROWS_8COL)
    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())
    assert all(item.values["warehouse"] for item in report.all_rows)
    # опциональные поля отсутствуют в формате → None
    assert all(item.values["markdown"] is None for item in report.all_rows)
    assert all(item.values["supplier"] is None for item in report.all_rows)


def test_two_unheaded_text_columns_is_error(tmp_path):
    rows = [[1, "ART", "Деталь", "VAG", 100, 200, 1, "Склад", "мусор-текст"]]
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, header in enumerate(HEADERS_8COL + [None], start=1):
        ws.cell(2, col, header)
    for col, value in enumerate(rows[0], start=1):
        ws.cell(3, col, value)
    inp = tmp_path / "in.xlsx"
    wb.save(inp)
    with pytest.raises(ValueError, match="Склад/Цех"):
        reprice_file(inp, tmp_path / "out.xlsx", Rules())


def test_missing_required_article_is_error(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, header in enumerate(["Что-то", "Цена", "Min Цена", "Склад/Цех"], start=1):
        ws.cell(1, col, header)
    inp = tmp_path / "in.xlsx"
    wb.save(inp)
    # нет «Артикул» → строка заголовка вообще не находится
    with pytest.raises(ValueError, match="Не найден лист"):
        reprice_file(inp, tmp_path / "out.xlsx", Rules())


def test_header_found_in_first_5_rows_only(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, header in enumerate(HEADERS_8COL, start=1):
        ws.cell(6, col, header)  # 6-я строка — за пределами поиска
    inp = tmp_path / "in.xlsx"
    wb.save(inp)
    with pytest.raises(ValueError, match="Не найден лист"):
        reprice_file(inp, tmp_path / "out.xlsx", Rules())


def test_markdown_synonym_and_operator_column_untouched(tmp_path):
    # 15-колоночный вариант: «Уценка согл.» = уценка, «Нов цена-4%» — колонка
    # оператора, которую не трогаем; склад без заголовка в конце
    headers = [None, "Артикул", "Номенклатура", "Бренд", "Цена", "Min Цена",
               "Поставщик", "Количество", "Уценка согл.", "Нов цена-4%", None]
    rows = [
        # база 960 < себест 2000, уценка 500 → 500*1.05 = 525
        [1, "X1", "Деталь", "VAG", 2000, 1000, "site.ru", 2, 500, 777.77, "Склад А"],
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Проценка"
    for col, header in enumerate(headers, start=1):
        ws.cell(2, col, header)
    for col, value in enumerate(rows[0], start=1):
        ws.cell(3, col, value)
    inp = tmp_path / "in.xlsx"
    wb.save(inp)

    report = reprice_file(inp, tmp_path / "out.xlsx", Rules())
    assert report.by_status == {"УЦЕНКА+5%": 1}

    out = openpyxl.load_workbook(tmp_path / "out.xlsx")["Проценка"]
    assert out.cell(3, 10).value == 777.77          # операторская «Нов цена-4%» не тронута
    assert out.cell(3, 11).value == "Склад А"        # склад по фолбэку
    assert out.cell(2, 12).value == "Новая цена"     # наши колонки дописаны после
    assert out.cell(3, 12).value == 525
    assert out.cell(2, 14).value == "Статус"


# --- Приёмка: все реальные файлы парсятся одним кодом, без флагов ---

@pytest.mark.parametrize(
    "fixture,expected_rows,has_markdown",
    [
        ("procenka_13col_2026-06-09.xlsx", 1180, True),
        ("procenka_8col_2026-07-06.xlsx", 1205, False),
        ("procenka_15col_2026-07-06_markdown.xlsx", 1205, True),
    ],
)
def test_real_files_single_parser(tmp_path, fixture, expected_rows, has_markdown):
    src = FIXTURES / fixture
    if not src.is_file():
        pytest.skip(f"нет фикстуры {fixture}")

    out = tmp_path / "out.xlsx"
    report = reprice_file(src, out, Rules())

    assert report.total == expected_rows
    assert sum(report.by_status.values()) == expected_rows
    # склад определён у каждой строки
    assert all(item.values["warehouse"] for item in report.all_rows)
    # уценка: в 13-колоночном есть у части строк, в 8-колоночном отсутствует
    markdown_rows = sum(
        1 for item in report.all_rows
        if isinstance(item.values["markdown"], (int, float))
    )
    assert (markdown_rows > 0) == has_markdown

    # выход открывается и содержит статусы
    ws = openpyxl.load_workbook(out)["Проценка"]
    assert ws.max_row >= expected_rows
